#!/usr/bin/env python3
import argparse
import csv
import json
import os
from pathlib import Path
from statistics import fmean, pstdev
from urllib import error, request

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = REPO_ROOT / "outputs" / "vw_passat_2016_raw_route_classified_60d_10s.xlsx"
DEFAULT_OUTPUT_DIR = REPO_ROOT / "outputs" / "datawrapper_city_only"
DEFAULT_WORKSPACE = "qeagfnuuhp"
TOKEN_ENV = "DATAWRAPPER_TOKEN"
API_BASE = "https://api.datawrapper.de/v3"
CITY_SHEET = "City_Only"
BIN_MINUTES = 20
INTRO_TEXT = "Series agregadas cada 20 minutos."
NOTES_TEXT = "Solo incluye trayectos clasificados como city."
CORE_METRICS = [
    "speed_kmh",
    "rpm",
    "coolant_c",
    "engine_load_pct",
    "oil_temp_c",
    "runtime_s",
    "throttle_pct",
    "intake_air_c",
    "maf_gps",
    "fuel_rate_lph",
]
METRIC_LABELS = {
    "speed_kmh": "Speed (km/h)",
    "rpm": "RPM",
    "coolant_c": "Coolant (C)",
    "engine_load_pct": "Engine Load (%)",
    "oil_temp_c": "Oil Temp (C)",
    "runtime_min": "Runtime (min)",
    "throttle_pct": "Throttle (%)",
    "intake_air_c": "Intake Air (C)",
    "maf_gps": "MAF (g/s)",
    "fuel_rate_lph": "Fuel Rate (L/h)",
}
DATA_HEADERS = [
    "timestamp",
    "speed_kmh",
    "rpm",
    "coolant_c",
    "engine_load_pct",
    "oil_temp_c",
    "runtime_min",
    "throttle_pct",
    "intake_air_c",
    "maf_gps",
    "fuel_rate_lph",
]
Z_HEADERS = [
    "timestamp",
    "z_speed_kmh",
    "z_rpm",
    "z_coolant_c",
    "z_engine_load_pct",
    "z_oil_temp_c",
    "z_runtime_min",
    "z_throttle_pct",
    "z_intake_air_c",
    "z_maf_gps",
    "z_fuel_rate_lph",
]
OVERLAY_CHART_SPEC = {
    "title": "VW Passat 2016 - City Only - Z-Score Overlay",
    "type": "d3-lines",
    "csv_name": "city_only_core_20min_zscore.csv",
}


def floor_to_bin(timestamp):
    minute = (timestamp.minute // BIN_MINUTES) * BIN_MINUTES
    return timestamp.replace(minute=minute, second=0, microsecond=0)


def blank_row(headers: list[str]) -> dict[str, str]:
    return {header: "" for header in headers}


def flush_bucket(rows: list[dict[str, object]], bucket: dict[str, object] | None) -> None:
    if not bucket:
        return

    row = {"timestamp": bucket["bucket_start"].strftime("%Y-%m-%d %H:%M:%S")}
    for metric in CORE_METRICS:
        if metric == "runtime_s":
            row["runtime_min"] = round(bucket["runtime_s_max"] / 60.0, 4)
            continue

        total = bucket[f"{metric}_sum"]
        count = bucket[f"{metric}_count"]
        row[metric] = round(total / count, 4) if count else ""
    rows.append(row)


def build_bucket(bucket_start):
    bucket = {"bucket_start": bucket_start, "runtime_s_max": 0.0}
    for metric in CORE_METRICS:
        if metric == "runtime_s":
            continue
        bucket[f"{metric}_sum"] = 0.0
        bucket[f"{metric}_count"] = 0
    return bucket


def aggregate_city_rows(input_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook[CITY_SHEET]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    index = {name: position for position, name in enumerate(header_row)}

    rows: list[dict[str, object]] = []
    current_segment = None
    current_bucket_start = None
    bucket = None

    for values in sheet.iter_rows(min_row=2, values_only=True):
        timestamp = values[index["timestamp"]]
        segment_id = values[index["segment_id"]]
        bucket_start = floor_to_bin(timestamp)

        if current_segment is None:
            current_segment = segment_id
            current_bucket_start = bucket_start
            bucket = build_bucket(bucket_start)
        elif segment_id != current_segment:
            flush_bucket(rows, bucket)
            rows.append(blank_row(DATA_HEADERS))
            current_segment = segment_id
            current_bucket_start = bucket_start
            bucket = build_bucket(bucket_start)
        elif bucket_start != current_bucket_start:
            flush_bucket(rows, bucket)
            current_bucket_start = bucket_start
            bucket = build_bucket(bucket_start)

        runtime_s = values[index["runtime_s"]]
        if runtime_s is not None and runtime_s > bucket["runtime_s_max"]:
            bucket["runtime_s_max"] = float(runtime_s)

        for metric in CORE_METRICS:
            if metric == "runtime_s":
                continue
            value = values[index[metric]]
            if value is None:
                continue
            bucket[f"{metric}_sum"] += float(value)
            bucket[f"{metric}_count"] += 1

    flush_bucket(rows, bucket)
    if rows and rows[-1]["timestamp"] == "":
        rows.pop()
    return rows


def write_csv(rows: list[dict[str, object]], headers: list[str], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def build_single_metric_rows(rows: list[dict[str, object]], metric_header: str) -> list[dict[str, object]]:
    metric_rows = []
    for row in rows:
        if not row["timestamp"]:
            metric_rows.append({"timestamp": "", metric_header: ""})
            continue
        metric_rows.append({"timestamp": row["timestamp"], metric_header: row[metric_header]})
    return metric_rows


def build_zscore_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    series = {header: [] for header in DATA_HEADERS if header != "timestamp"}
    for row in rows:
        if not row["timestamp"]:
            continue
        for header in series:
            value = row[header]
            if value != "":
                series[header].append(float(value))

    stats = {}
    for header, values in series.items():
        if not values:
            stats[header] = (0.0, 0.0)
            continue
        stats[header] = (fmean(values), pstdev(values))

    z_rows = []
    for row in rows:
        if not row["timestamp"]:
            z_rows.append(blank_row(Z_HEADERS))
            continue

        z_row = {"timestamp": row["timestamp"]}
        for header in DATA_HEADERS[1:]:
            mean, stddev = stats[header]
            value = float(row[header])
            z_header = f"z_{header}"
            z_row[z_header] = 0 if stddev == 0 else round((value - mean) / stddev, 6)
        z_rows.append(z_row)
    return z_rows


def validate_rows(rows: list[dict[str, object]], headers: list[str], expected_min_rows: int) -> None:
    if len(headers) != len(set(headers)):
        raise SystemExit("duplicate CSV headers")
    if len(rows) < expected_min_rows:
        raise SystemExit(f"expected at least {expected_min_rows} rows, got {len(rows)}")

    for row in rows:
        if row[headers[0]] == "":
            continue
        for header in headers[1:]:
            value = row[header]
            if value == "":
                continue
            number = float(value)
            if number != number or number in (float("inf"), float("-inf")):
                raise SystemExit(f"invalid numeric value in {header}: {value}")


def api_request(method: str, path: str, token: str, payload=None, headers=None):
    request_headers = {"Authorization": f"Bearer {token}", "accept": "application/json"}
    if headers:
        request_headers.update(headers)

    data = None
    if payload is not None:
        if isinstance(payload, str):
            data = payload.encode("utf-8")
        elif isinstance(payload, bytes):
            data = payload
        else:
            data = json.dumps(payload).encode("utf-8")
            request_headers.setdefault("content-type", "application/json")

    url = f"{API_BASE}{path}"
    req = request.Request(url, method=method, data=data, headers=request_headers)
    try:
        with request.urlopen(req) as response:
            body = response.read().decode("utf-8")
            content_type = response.headers.get("content-type", "")
            if "application/json" in content_type:
                return json.loads(body)
            if not body:
                return None
            return body
    except error.HTTPError as exc:
        body = exc.read().decode("utf-8")
        raise SystemExit(f"{method} {path} failed with {exc.code}: {body}") from exc


def create_chart(token: str, workspace: str, title: str, chart_type: str) -> dict[str, object]:
    return api_request(
        "POST",
        "/charts",
        token,
        payload={"title": title, "type": chart_type, "workspace": workspace},
    )


def upload_chart_data(token: str, chart_id: str, csv_text: str) -> None:
    api_request(
        "PUT",
        f"/charts/{chart_id}/data",
        token,
        payload=csv_text,
        headers={"content-type": "text/csv"},
    )


def patch_chart_metadata(token: str, chart_id: str, title: str) -> dict[str, object]:
    payload = {
        "title": title,
        "language": "es-ES",
        "metadata": {
            "describe": {
                "intro": INTRO_TEXT,
            },
            "annotate": {
                "notes": NOTES_TEXT,
            },
        },
    }
    return api_request(
        "PATCH",
        f"/charts/{chart_id}",
        token,
        payload=payload,
        headers={"content-type": "application/merge-patch+json"},
    )


def get_chart_data_preview(token: str, chart_id: str) -> str:
    preview = api_request("GET", f"/charts/{chart_id}/data", token, headers={"accept": "text/csv"})
    if isinstance(preview, str):
        return "\n".join(preview.splitlines()[:3])
    raise SystemExit(f"unexpected data preview response for chart {chart_id}")


def get_editor_url(chart_id: str) -> str:
    return f"https://app.datawrapper.de/chart/{chart_id}/visualize"


def create_login_link(token: str, chart_id: str) -> str | None:
    try:
        response = api_request(
            "POST",
            "/auth/login-tokens",
            token,
            payload={"chartId": chart_id, "step": "visualize"},
        )
    except SystemExit as exc:
        if "POST /auth/login-tokens failed with 403" in str(exc):
            return None
        raise

    if isinstance(response, dict):
        if "url" in response:
            return response["url"]
        if "token" in response:
            return f"https://app.datawrapper.de/auth/login?token={response['token']}"
    raise SystemExit(f"unexpected login token response for chart {chart_id}: {response}")


def build_artifacts(input_path: Path, output_dir: Path) -> dict[str, object]:
    city_rows = aggregate_city_rows(input_path)
    z_rows = build_zscore_rows(city_rows)

    validate_rows(city_rows, DATA_HEADERS, expected_min_rows=900)
    validate_rows(z_rows, Z_HEADERS, expected_min_rows=900)

    city_csv = output_dir / "city_only_core_20min.csv"
    z_csv = output_dir / OVERLAY_CHART_SPEC["csv_name"]
    write_csv(city_rows, DATA_HEADERS, city_csv)
    write_csv(z_rows, Z_HEADERS, z_csv)

    metric_csvs = {}
    for metric_header in DATA_HEADERS[1:]:
        metric_rows = build_single_metric_rows(city_rows, metric_header)
        validate_rows(metric_rows, ["timestamp", metric_header], expected_min_rows=900)
        metric_csv = output_dir / f"{metric_header}_20min.csv"
        write_csv(metric_rows, ["timestamp", metric_header], metric_csv)
        metric_csvs[metric_header] = metric_csv

    separator_rows = sum(1 for row in city_rows if row["timestamp"] == "")
    return {
        "city_csv": city_csv,
        "z_csv": z_csv,
        "metric_csvs": metric_csvs,
        "row_count": len(city_rows),
        "separator_rows": separator_rows,
    }


def run(token: str, workspace: str, input_path: Path, output_dir: Path) -> dict[str, object]:
    me = api_request("GET", "/me", token)
    workspace_slugs = {item["slug"] for item in me.get("workspaces", [])}
    if workspace not in workspace_slugs:
        raise SystemExit(f"workspace '{workspace}' not found in token account")

    artifacts = build_artifacts(input_path, output_dir)
    results = {
        "artifacts": {
            "city_csv": str(artifacts["city_csv"]),
            "z_csv": str(artifacts["z_csv"]),
            "metric_csvs": {key: str(value) for key, value in artifacts["metric_csvs"].items()},
        }
    }
    results["row_count"] = artifacts["row_count"]
    results["separator_rows"] = artifacts["separator_rows"]
    results["charts"] = []

    for metric_header in DATA_HEADERS[1:]:
        title = f"VW Passat 2016 - City Only - {METRIC_LABELS[metric_header]}"
        chart = create_chart(token, workspace, title, "d3-lines")
        chart_id = chart["id"]
        csv_path = artifacts["metric_csvs"][metric_header]
        upload_chart_data(token, chart_id, csv_path.read_text())
        patched = patch_chart_metadata(token, chart_id, title)
        data_preview = get_chart_data_preview(token, chart_id)
        edit_url = get_editor_url(chart_id)
        login_url = create_login_link(token, chart_id)

        if patched["type"] != "d3-lines":
            raise SystemExit(f"chart {chart_id} has unexpected type {patched['type']}")

        results["charts"].append(
            {
                "metric": metric_header,
                "id": chart_id,
                "title": title,
                "type": "d3-lines",
                "edit_url": edit_url,
                "login_url": login_url,
                "data_preview": data_preview,
            }
        )

    overlay_chart = create_chart(token, workspace, OVERLAY_CHART_SPEC["title"], OVERLAY_CHART_SPEC["type"])
    overlay_chart_id = overlay_chart["id"]
    upload_chart_data(token, overlay_chart_id, artifacts["z_csv"].read_text())
    patched = patch_chart_metadata(token, overlay_chart_id, OVERLAY_CHART_SPEC["title"])
    data_preview = get_chart_data_preview(token, overlay_chart_id)
    edit_url = get_editor_url(overlay_chart_id)
    login_url = create_login_link(token, overlay_chart_id)

    if patched["type"] != OVERLAY_CHART_SPEC["type"]:
        raise SystemExit(f"chart {overlay_chart_id} has unexpected type {patched['type']}")

    results["overlay_chart"] = {
        "id": overlay_chart_id,
        "title": OVERLAY_CHART_SPEC["title"],
        "type": OVERLAY_CHART_SPEC["type"],
        "edit_url": edit_url,
        "login_url": login_url,
        "data_preview": data_preview,
    }

    return results


def main() -> int:
    parser = argparse.ArgumentParser(description="Create city-only Datawrapper charts from AutoSense telemetry.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Route-classified XLSX input")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="CSV artifact output directory")
    parser.add_argument("--workspace", default=DEFAULT_WORKSPACE, help="Datawrapper workspace slug")
    args = parser.parse_args()

    token = os.environ.get(TOKEN_ENV)
    if not token:
        raise SystemExit(f"missing {TOKEN_ENV}")
    if not args.input.exists():
        raise SystemExit(f"input file not found: {args.input}")

    results = run(token, args.workspace, args.input, args.output_dir)
    results_path = args.output_dir / "datawrapper_city_only_result.json"
    results_path.write_text(json.dumps(results, indent=2))
    print(json.dumps(results, indent=2))
    print(f"saved results to {results_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
